import os
import zipfile
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from makedoc.utils import (
    get_formatted_date_agreement,
    get_formatted_date_shipment,
)
from .data_service import DataService


class Documents:
    """
    A class to handle the creation and management of various shipment-related documents.
    """

    def __init__(self, validated_data):
        self.validated_data = validated_data
        self.auto = 0
        self.rw = 0
        self.service_note = 0
        self.transport_sheet = 0
        self.archive_name = ""

    def update_documents(self):
        """
        Updates the document indicators based on the validated data.
        """
        delivery_type = DataService.get_delivery_type(self.validated_data)
        if delivery_type in ("auto", "self-delivery"):
            self.auto = 1
        if delivery_type == "rw":
            self.rw = 1

        if self.validated_data["delivery_cost"] > 0:
            self.transport_sheet = 1

        have_discount = max(self.validated_data.get("items"), key=lambda x: x["discount"])[
            "discount"
        ]

        if have_discount > 0:
            self.service_note = 1

    def form_auto_document(self, request):
        """
        Creates an auto document based on the validated data and user request.
        """
        self.docname = "Авто"
        try:
            user = DataService.get_user(request)
            client = DataService.get_client(self.validated_data)
            products = DataService.get_products(self.validated_data)
            factory = DataService.get_factory(self.validated_data)
            logistics = DataService.get_delivery_cost(self.validated_data)
        except Exception as e:
            return f"Error fetching data: {e}"

        template_path = "makedoc/excel-templates/spec.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        self.fill_contract_info(ws, client)
        self.fill_product_info(ws, products, logistics, 10)
        self.fill_factory_info(ws, factory)
        self.fill_auto_services(ws, logistics)
        self.fill_debt_info(ws)
        self.fill_legal_info(ws, client)
        self.fill_signatures(ws, client)
        self.fill_manager_contact(ws, user)

        self.apply_styles(ws)

        self.add_workbook_to_archive_xlsx(wb, user, client)

    def fill_contract_info(self, ws, client):
        """
        Fills the contract information into the worksheet.
        """
        formatted_contract_date = client.contract_date.strftime("%d.%m.%Y")
        title = ws.cell(row=1, column=1, value=f"Приложение № {client.last_application_number}")
        title.font = Font(bold=True)
        ws.cell(
            row=2,
            column=1,
            value=f"к договору поставки № {client.contract_number} от {formatted_contract_date}г.",
        )
        ws.cell(row=4, column=1, value=client.client_name)
        ws.cell(row=6, column=6, value=get_formatted_date_agreement())

    def fill_product_info(self, ws, products, logistics, caret):
        """
        Fills the product information into the worksheet.
        """
        goods_quantity = len(products)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for i in range(goods_quantity):
            ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=3)
            # Название продукта
            product = f"{products[i]['product'].flour_name.flour_name}, \
т/м {products[i]['product'].brand.brand}"
            ws.cell(row=caret, column=1, value=product)
            # Упаковка
            package = products[i]["product"].package.package
            ws.cell(row=caret, column=4, value=package)
            # Количество товара по заказу
            ws.cell(row=caret, column=5, value=products[i]["quantity"])
            # Цена за товар
            price = products[i]["price"] * (100 - products[i]["discount"]) / 100
            ws.cell(row=caret, column=6, value=price)
            if logistics:
                # Стоимость доставки
                ws.cell(row=caret, column=7, value=logistics)
                # Цена на самовывозе
                ws.cell(row=caret, column=8, value=price - logistics)

            table_width = range(1, 9)
            if not logistics:
                table_width = range(1, 7)

            for col_num in table_width:
                cell = ws.cell(row=caret, column=col_num)
                cell.border = thin_border
                if cell.column > 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            caret += 1
        self.caret_product = caret + 1

    def fill_factory_info(self, ws, factory):
        """
        Fills the factory information into the worksheet.
        """
        caret = self.caret_product
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=2)
        ws.cell(row=caret, column=1, value="Грузоотправитель:")
        ws.cell(row=caret, column=3, value=factory.full_name)
        self.caret_factory = caret + 1

    def fill_auto_services(self, ws, logistics):
        """
        Fills the auto services information into the worksheet.
        """
        caret = self.caret_factory
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=2)
        ws.cell(row=caret, column=1, value="Автотранспортные услуги:")
        ws.cell(
            row=caret,
            column=3,
            value="не входят в стоимость товара" if not logistics else "входят в стоимость товара",
        )
        self.caret_services = caret + 1

    def fill_debt_info(self, ws):
        """
        Fills the debt information into the worksheet.
        """
        caret = self.caret_services
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=2)
        ws.cell(row=caret, column=1, value="Срок отгрузки:")
        ws.cell(row=caret, column=3, value=get_formatted_date_shipment("nomn"))
        ws.merge_cells(start_row=caret + 1, start_column=1, end_row=caret + 1, end_column=6)
        pdz = "▪Продавец имеет право не осуществлять отгрузку товара до полного погашения \
Покупателем просроченной дебиторской задолженности."
        debt = ws.cell(row=caret + 1, column=1, value=pdz)
        debt.alignment = Alignment(wrap_text=True, horizontal="justify", vertical="center")
        ws.row_dimensions[caret + 1].height = 30
        self.caret_debt = caret + 2

    def fill_legal_info(self, ws, client):
        """
        Fills the legal information into the worksheet.
        """
        caret = self.caret_debt
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=6)
        formatted_contract_date = client.contract_date.strftime("%d.%m.%Y")
        contract_option = f"▪Настоящее приложение составлено и подписано в двух экземплярах, \
имеющих одинаковую юридическую силу, по одному для каждой из сторон, вступает в силу с момента \
подписания и является неотъемлемой частью договора № {client.contract_number} от \
{formatted_contract_date}г."
        legal = ws.cell(row=caret, column=1, value=contract_option)
        legal.alignment = Alignment(wrap_text=True, horizontal="justify", vertical="center")
        ws.row_dimensions[caret].height = 50
        self.caret_legal = caret + 4

    def fill_signatures(self, ws, client):
        """
        Fills the signature information into the worksheet.
        """
        caret = self.caret_legal
        ws.cell(row=caret, column=1, value="Генеральный директор")
        ws.cell(row=caret + 1, column=1, value="ООО  «Торговый дом «Оскольская мука»")
        ws.cell(row=caret + 1, column=6, value="С.А. Годизов")
        ws.cell(row=caret + 9, column=1, value=str(client.director_position))
        ws.cell(row=caret + 10, column=1, value=client.client_name)
        ws.cell(row=caret + 10, column=6, value=client.director_name)
        self.caret_signatures = caret + 11

    def fill_manager_contact(self, ws, user):
        """
        Fills the manager contact information into the worksheet.
        """
        caret = 59
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=6)
        manager = ws.cell(
            row=caret, column=1, value=f"Ваш персональный менеджер: {user.full_name}"
        )
        manager.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=caret + 1, start_column=1, end_row=caret + 1, end_column=6)
        phone = ws.cell(
            row=caret + 1,
            column=1,
            value=f"Тел. {user.phone_number_personal}, моб. {user.phone_number_work}",
        )
        phone.alignment = Alignment(horizontal="center", vertical="center")

    def add_workbook_to_archive_xlsx(self, wb, user, client):
        """
        Adds the generated workbook to a zip archive.
        """
        tempdir = os.path.join("makedoc", "tempdoc", str(user.id))
        os.makedirs(tempdir, exist_ok=True)

        client_name = client.client_name.replace(" ", "_")
        date_today = datetime.today().strftime("%d.%m.%Y_%H:%M:%S")
        self.archive_name = f"{client_name}_{date_today}.zip"
        archive_path = f"{tempdir}/{self.archive_name}"

        xlsx_filename = f"{client.last_application_number} {self.docname} {client.client_name} \
{date.today().strftime('%d.%m.%Y')}.xlsx"

        xlsx_io = BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        with zipfile.ZipFile(archive_path, "a") as archive:
            archive.writestr(xlsx_filename, xlsx_io.getvalue())

    def apply_styles(self, ws):
        """
        Applies styles to the cells in the worksheet.
        """
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(name="Times New Roman", size=12)

        for row_num in [1, 9]:
            for cell in ws[row_num]:
                if cell.value is not None:
                    cell.font = Font(bold=True, size=12)

    def form_rw_document(self, request):
        """
        Creates a railways document based on the validated data and user request.
        """
        self.docname = "Жд"
        try:
            user = DataService.get_user(request)
            client = DataService.get_client(self.validated_data)
            product = DataService.get_products(self.validated_data)
            factory = DataService.get_factory(self.validated_data)
            logistics = DataService.get_delivery_cost(self.validated_data)

        except Exception as e:
            return f"Error fetching data: {e}"

        template_path = "makedoc/excel-templates/spec.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        self.fill_contract_info(ws, client)
        self.fill_product_info(ws, product, logistics, 10)
        self.fill_factory_info(ws, factory)
        self.fill_rw_services(ws, factory, client, logistics)
        self.fill_debt_info(ws)
        self.fill_legal_info(ws, client)
        self.fill_signatures(ws, client)
        self.fill_manager_contact(ws, user)

        self.apply_styles(ws)

        self.add_workbook_to_archive_xlsx(wb, user, client)

    def fill_rw_services(self, ws, factory, client, logistics):
        """
        Fills the railways services information into the worksheet.
        """
        caret = self.caret_factory
        ws.merge_cells(start_row=caret, start_column=1, end_row=caret, end_column=2)
        ws.cell(row=caret, column=1, value="Поставка осуществляется:")
        ws.cell(row=caret, column=3, value="ж/д транспортом")
        caret += 1

        ws.cell(row=caret, column=1, value="Станция отправления:")
        shipping_station = f"{factory.departure_station_name}, {factory.departure_station_branch}"
        ws.cell(row=caret, column=3, value=shipping_station)
        caret += 1

        ws.cell(row=caret, column=1, value="Условия поставки:")
        ws.cell(
            row=caret,
            column=3,
            value="франко-вагон станция отправления"
            if not logistics
            else "франко-вагон станция назначения",
        )
        caret += 2

        ws.cell(row=caret, column=1, value="Отгрузочные реквизиты:")
        caret += 1

        ws.cell(row=caret, column=1, value="Станция назначения:")
        destination_station = (
            f"{client.railway_station.station_name}, {client.railway_station.station_branch}"
        )
        ws.cell(row=caret, column=3, value=destination_station)
        caret += 1

        ws.cell(row=caret, column=1, value="Код станции:")
        station_code_value = ws.cell(row=caret, column=3, value=client.railway_station.station_id)
        station_code_value.alignment = Alignment(horizontal="left")
        caret += 1

        ws.cell(row=caret, column=1, value="Получатель:")
        ws.cell(row=caret, column=3, value=client.receiver_name)
        caret += 1

        ws.cell(row=caret, column=1, value="Код получателя:")
        receiver_code_value = ws.cell(row=caret, column=3, value=client.receiver_id)
        receiver_code_value.alignment = Alignment(horizontal="left")
        caret += 1

        ws.cell(row=caret, column=1, value="ОКПО:")
        okpo_value = ws.cell(row=caret, column=3, value=client.receiver_okpo)
        okpo_value.alignment = Alignment(horizontal="left")
        caret += 1

        ws.cell(row=caret, column=1, value="Адрес:")
        ws.cell(row=caret, column=3, value=client.receiver_adress)
        caret += 1

        ws.cell(row=caret, column=1, value="Особые отметки:")
        ws.cell(row=caret, column=3, value=client.special_marks)
        caret += 1

        self.caret_services = caret + 1

    def form_service_note(self, request):
        """
        Creates a service note based on the validated data and user request.
        """
        self.docname = "Служебная записка"
        try:
            client = DataService.get_client(self.validated_data)
            destination = DataService.get_destination(self.validated_data)
            user = DataService.get_user(request)
            product = DataService.get_products(self.validated_data)
            discount = max(product, key=lambda x: x["discount"])["discount"]

            logistics = DataService.get_delivery_cost(self.validated_data)
        except Exception as e:
            return f"Error fetching data: {e}"

        # Открытие файла шаблона
        template_path = "makedoc/excel-templates/service_note.xlsx"
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active

        self.fill_text_note(ws, client, discount, destination)
        self.fill_product_info(ws, product, logistics, 22)
        self.apply_styles(ws)

        self.add_workbook_to_archive_xlsx(wb, user, client)

    def fill_text_note(self, ws, client, discount, destination):
        """
        Fills the text note section in the service note Excel sheet.
        """
        blank = "________"
        if destination != "0":
            region = destination.split(",")[1].strip()
        else:
            destination, region = blank, blank
        ws.cell(row=15, column=1, value=f"{get_formatted_date_agreement()} № 12/2.2/23/3-___")
        # Нужно правильлно разобрать в json регион чтобы склонять только название
        text = f"    В целях увеличения объема продаж на территории {region} прошу Вашего \
согласования применить скидку для контрагента {client.client_name} ({destination}) до {discount}%\
 в {get_formatted_date_shipment('loct')} на продукцию следующего ассортимента: "
        ws.cell(row=19, column=1, value=text)

    def form_transport_sheet(self, request):
        """
        Creates a transport sheet document based on the validated data and user request.
        """
        self.docname = "Сопроводительный лист"
        try:
            user = DataService.get_user(request)
            product = DataService.get_products(self.validated_data)
            client = DataService.get_client(self.validated_data)
            logistics = DataService.get_delivery_cost(self.validated_data)
            factory = DataService.get_factory(self.validated_data)

        except Exception as e:
            return f"Error fetching data: {e}"

        template_path = "makedoc/excel-templates/transport_sheet.xlsx"
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active

        self.fill_contract_info_transport_sheet(ws, client)
        self.fill_product_info(ws, product, logistics, 11)

        self.fill_factory_info(ws, factory)
        self.fill_auto_services(ws, logistics)
        self.fill_debt_info(ws)
        self.fill_legal_info(ws, client)
        self.fill_signatures(ws, client)
        self.fill_manager_contact(ws, user)

        self.apply_styles(ws)

        self.add_workbook_to_archive_xlsx(wb, user, client)

    def fill_contract_info_transport_sheet(self, ws, client):
        """
        Fills the contract information section in the transport sheet Excel sheet.
        """
        formatted_contract_date = client.contract_date.strftime("%d.%m.%Y")
        ws.cell(row=1, column=1, value="СОПРОВОДИТЕЛЬНЫЙ ЛИСТ к")
        title = ws.cell(row=2, column=1, value=f"Приложению № {client.last_application_number}")
        title.font = Font(bold=True)
        ws.cell(
            row=3,
            column=1,
            value=f"к договору поставки № {client.contract_number} от {formatted_contract_date}г.",
        )
        ws.cell(row=5, column=1, value=client.client_name)
        ws.cell(row=7, column=8, value=get_formatted_date_agreement())
