import datetime
import openpyxl

from django.http import HttpResponse
from dateutil.relativedelta import relativedelta
from .services import get_client, get_user, get_client_rw, get_rw, get_logistics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from .constants import MONTHS_AGREEMENT, MONTHS_SHIPMENT


def write_to_excel_auto(request):
    client = get_client(request)
    user = get_user(request)
    # Открытие файла шаблона
    template_path = "exel-templates/auto.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Форматирование даты договора
    formatted_contract_date = client.contract_date.strftime("%d.%m.%Y")

    # Сегодняшняя дата
    current_date = datetime.datetime.today()
    formatted_date_agreement = f'«{current_date.day}» {MONTHS_AGREEMENT[current_date.strftime("%B")]} {current_date.year} г.'

    # Отгрузка
    next_month_date = current_date + relativedelta(months=+1)
    formatted_date_shipment = f"{MONTHS_SHIPMENT[current_date.strftime('%B')]}-{MONTHS_SHIPMENT[next_month_date.strftime('%B')]} {current_date.year} г."

    # Разьеденить ячейки
    worksheet.unmerge_cells("A1:F1")
    worksheet.unmerge_cells("A2:F2")
    worksheet.unmerge_cells("A4:F4")
    worksheet.unmerge_cells("A17:F17")
    worksheet.unmerge_cells("A49:F49")
    worksheet.unmerge_cells("A50:F50")

    # Записать значений в ячейки
    worksheet["A1"] = f"Приложение № {client.last_application_number}"
    worksheet["A2"] = (
        f"к договору поставки № {client.contract_number} от {formatted_contract_date}г."
    )
    worksheet["A4"] = f"ООО  (ИП, АО)  «{client.client_name}»"
    worksheet["F6"] = formatted_date_agreement
    worksheet["A17"] = (
        f"▪Настоящее приложение составлено и подписано в двух экземплярах, имеющих одинаковую юридическую силу, по одному для каждой из сторон, вступает в силу с момента подписания и является неотъемлемой частью договора № {client.contract_number} от {formatted_contract_date}г."
    )
    worksheet["C14"] = formatted_date_shipment
    worksheet["A35"] = f"{client.director_position}"
    worksheet["A36"] = f"{client.client_name}"
    worksheet["F36"] = f"{client.director_name}"
    worksheet["A49"] = f"Ваш персональный менеджер: {user.full_name}"
    worksheet["A50"] = f"Тел. {user.phone_number_personal}, моб. {user.phone_number_work}"

    # Объединить ячейки
    worksheet.merge_cells("A1:F1")
    worksheet.merge_cells("A2:F2")
    worksheet.merge_cells("A4:F4")
    worksheet.merge_cells("A17:F17")
    worksheet.merge_cells("A49:F49")
    worksheet.merge_cells("A50:F50")

    # Сохранить файл
    new_file_path = f"auto_{user.full_name.split()[0]}_{current_date.strftime('%d.%m.%Y')}.xlsx"
    workbook.save(new_file_path)

    return HttpResponse(f"Документ сохранен как {new_file_path}")


def write_to_excel_rw(request):
    client = get_client_rw(request)
    user = get_user(request)
    rw = get_rw(request)
    # Открытие файла шаблона
    template_path = "exel-templates/rw.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Форматирование даты договора
    formatted_contract_date = client.contract_date.strftime("%d.%m.%Y")

    # Сегодняшняя дата
    current_date = datetime.datetime.today()
    formatted_date_agreement = f'«{current_date.day}» {MONTHS_AGREEMENT[current_date.strftime("%B")]} {current_date.year} г.'

    # Отгрузка
    next_month_date = current_date + relativedelta(months=+1)
    formatted_date_shipment = f"{MONTHS_SHIPMENT[current_date.strftime('%B')]}-{MONTHS_SHIPMENT[next_month_date.strftime('%B')]} {current_date.year} г."

    # Разьеденить ячейки
    worksheet.unmerge_cells("A1:F1")
    worksheet.unmerge_cells("A2:F2")
    worksheet.unmerge_cells("A4:F4")
    worksheet.unmerge_cells("A26:F26")
    worksheet.unmerge_cells("A49:F49")
    worksheet.unmerge_cells("A50:F50")

    # Записать значений в ячейки
    worksheet["A1"] = f"Приложение № {client.last_application_number}"
    worksheet["A2"] = (
        f"к договору поставки № {client.contract_number} от {formatted_contract_date}г."
    )
    worksheet["A4"] = f"ООО  (ИП, АО)  «{client.client_name}»"
    worksheet["F6"] = formatted_date_agreement
    worksheet["A26"] = (
        f"▪Настоящее приложение составлено и подписано в двух экземплярах, имеющих одинаковую юридическую силу, по одному для каждой из сторон, вступает в силу с момента подписания и является неотъемлемой частью договора № {client.contract_number} от {formatted_contract_date}г."
    )
    worksheet["C16"] = formatted_date_shipment
    worksheet["C19"] = rw.station_name
    worksheet["C20"] = rw.station_id
    worksheet["C21"] = f"ООО  (ИП, АО) {client.receiver_name}"
    worksheet["C22"] = f"{client.receiver_id}"
    worksheet["C23"] = f"{client.receiver_okpo}"
    worksheet["C24"] = f"{client.receiver_adress}"
    worksheet["A42"] = f"{client.director_position}"
    worksheet["A43"] = f"{client.client_name}"
    worksheet["F43"] = f"{client.director_name}"
    worksheet["A49"] = f"Ваш персональный менеджер: {user.full_name}"
    worksheet["A50"] = f"Тел. {user.phone_number_personal}, моб. {user.phone_number_work}"

    # Объединить ячейки
    worksheet.merge_cells("A1:F1")
    worksheet.merge_cells("A2:F2")
    worksheet.merge_cells("A4:F4")
    worksheet.merge_cells("A26:F26")
    worksheet.merge_cells("A49:F49")
    worksheet.merge_cells("A50:F50")

    # Сохранить файл
    new_file_path = f"rw_{user.full_name.split()[0]}_{current_date.strftime('%d.%m.%Y')}.xlsx"
    workbook.save(new_file_path)

    return HttpResponse(f"Документ сохранен как {new_file_path}")


def write_to_excel_sluzebnyi(request):
    user = get_user(request)
    # logistics = get_logistics(request)
    # Открытие файла шаблона
    template_path = "exel-templates/sluz.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Сегодняшняя дата
    current_date = datetime.datetime.today()
    formatted_date_agreement = f'«{current_date.day}» {MONTHS_AGREEMENT[current_date.strftime("%B")]} {current_date.year} г.'

    # Разьеденить ячейки
    worksheet.unmerge_cells("A19:H19")

    # Записка
    worksheet["A15"] = f"{formatted_date_agreement} № 12/2.2/23/3-"
    worksheet["A19"] = f"..."

    # Объединить ячейки
    worksheet.merge_cells("A19:H19")

    # Сохранить файл
    new_file_path = f"sluz_{user.full_name.split()[0]}_{current_date.strftime('%d.%m.%Y')}.xlsx"
    workbook.save(new_file_path)

    return HttpResponse(f"Документ сохранен как {new_file_path}")


class HomeView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        content = {"message": "Test Authentication page using React Js and Django!"}
        return Response(content)
