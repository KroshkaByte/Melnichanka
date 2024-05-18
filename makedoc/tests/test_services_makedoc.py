import os
import sys

import pytest

from makedoc.services import get_client, get_product, get_user, write_to_excel_auto

sys.path.append("path/to/your/module/directory")  # Добавьте путь к вашему модулю
from unittest.mock import MagicMock

import openpyxl


@pytest.fixture
def mock_request():
    request = MagicMock()
    request.user = MagicMock()
    request.client = MagicMock()
    request.product = MagicMock()
    return request


@pytest.fixture
def mock_user():
    return MagicMock(
        full_name="Test User", phone_number_personal="1234567890", phone_number_work="0987654321"
    )


@pytest.fixture
def mock_client():
    client = MagicMock()
    client.contract_number = "1234567890"
    client.contract_date = "2022-01-01"
    client.client_name = "Test Client"
    client.director_name = "John Doe"
    client.director_position = "CEO"
    client.last_application_number = 123
    return client


@pytest.fixture
def mock_product():
    product = MagicMock()
    product.flour_name = "Test Flour"
    product.brand = "Test Brand"
    product.package = MagicMock(package="Test Package")
    product.price = 10.0
    return product


def test_write_to_excel_auto_creates_file(
    mock_request, mock_user, mock_client, mock_product, tmp_path
):
    get_user.return_value = mock_user
    get_client.return_value = mock_client
    get_product.return_value = mock_product

    write_to_excel_auto(mock_request)

    assert os.path.exists(
        os.path.join(tmp_path, "makedoc", "01.01.2022", "Test", "auto_Test_01.01.2022.xlsx")
    )


def test_write_to_excel_auto_fills_data(
    mock_request, mock_user, mock_client, mock_product, tmp_path
):
    get_user.return_value = mock_user
    get_client.return_value = mock_client
    get_product.return_value = mock_product

    write_to_excel_auto(mock_request)

    wb = openpyxl.load_workbook(
        os.path.join(tmp_path, "makedoc", "01.01.2022", "Test", "auto_Test_01.01.2022.xlsx")
    )
    ws = wb.active

    assert ws.cell(row=1, column=1).value == f"Приложение № {mock_client.last_application_number}"
    assert (
        ws.cell(row=2, column=1).value
        == f"к договору поставки № {mock_client.contract_number} от 01.01.2022г."
    )
    assert ws.cell(row=4, column=1).value == mock_client.client_name
    assert ws.cell(row=6, column=6).value == "01.01.2022"
    assert ws.cell(row=10, column=1).value == f"{mock_product.flour_name} {mock_product.brand}"
    assert ws.cell(row=10, column=4).value == mock_product.package.package
    assert ws.cell(row=10, column=5).value == 20
    assert ws.cell(row=10, column=6).value == "10.0"


# Добавьте дополнительные тесты для проверки других частей файла Excel


@pytest.fixture(autouse=True)
def mock_get_formatted_date_agreement(mocker):
    mocker.patch("your_module.get_formatted_date_agreement", return_value="01.01.2022")


@pytest.fixture(autouse=True)
def mock_get_formatted_date_shipment(mocker):
    mocker.patch("your_module.get_formatted_date_shipment", return_value="01.01.2022")


@pytest.fixture(autouse=True)
def mock_factory(mocker):
    mocker.patch("your_module.Factory.objects.first", return_value="Test Factory")
